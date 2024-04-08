import os.path
import openpyxl
from slugify import slugify
import boto3
from boto3.exceptions import ClientError
from MersHub.settings import STATIC_ROOT, AWS_ACCESS_KEY_ID, \
    AWS_SECRET_ACCESS_KEY, AWS_STORAGE_BUCKET_NAME
from shop.models import Category, Product

category_dict = {
    'Капоты': 'Капот',
    'Петли капота': 'Капот',
    'Крылья': 'Крила',
    'Рейлинги': 'Кузов',
    'пластик': 'Кузов',
    'Подкрылки': 'Крила',
    'Ручки': 'Ручки',
    'Кронштейны ручек': 'Ручки',
    'Стекла': 'Стекла',
    'Резинки': 'Стекла',
    'Двери': 'Двері',
    '}{ром дверей': 'Двері',
    'Двери 1эт': 'Двері',
    'Двери  2эт.': 'Двері',
    'Багажнки': 'Багажник',
    'Бампера': 'Бампера',
    'Усилители бамперов': 'Бампера',
    'Пороги': 'Пороги',
    'Блоки розжига': 'Блоки',
    'Фары': 'Фари',
    'Фонари': 'Ліхтарі',
    'Проводка бампеов': 'Проводки',
    'Датчики': 'Датчики',
    'Лямбды': 'Датчики',
    'Радары': 'Датчики',
    'Кулаки': 'Кулаки',
    'Цапфы': 'Кулаки',
    'Рычаги': 'Важелі',
    'Педали': 'Педалі',
    'Подрамники': 'Підрамники',
    'Полуоси промвалы': 'Пів осі',
    'Карданы': 'Кардани',
    'АКПП': 'Коробки передач',
    'Аморты': 'Амортизатор',
    'Аморты капота': 'Амортизатор',
    'Суппорты': 'Супорти',
    'Редуктора': 'Редуктори',
    'Раздатки': 'Роздатки',
    'Моторы': 'Мотори',
    'Салоны': 'Салони',
    'Обшивка': 'Обшивка',
    'Рули': 'Кермо',
    'Рулевые колонки': 'Кермо',
    'Рейки': 'Кермо',
    'Безопасность': 'Безпека',
    'Подушки в руль': 'Безпека',
    'Защиты ДВС': 'Безпека',
    'Защиты днища': 'Безпека',
    'Замки разные': 'Замки',
    'Замки дверей': 'Замки',
    'Решетки': 'Інсталяційна панель (телевізори)',
    'Радиаторы': 'Вентилятор радіаторів'
}


def download_excel(filename):
    """
    Download an Excel file from an S3 bucket.

    This function downloads an Excel file from an S3 bucket using
    AWS credentials provided as environment variables. The file
    is saved locally to the specified location.

    Args:
        filename (str): The filename of the Excel file to download.

    Raises:
        Exception: If the download fails or AWS credentials are invalid.

    """
    try:
        # Initialize the S3 client
        s3 = boto3.client(
            's3',
            aws_access_key_id=AWS_ACCESS_KEY_ID,
            aws_secret_access_key=AWS_SECRET_ACCESS_KEY
        )
        # Download the file from S3 bucket
        s3.download_file(
            AWS_STORAGE_BUCKET_NAME,
            f'files/{filename}',
            f'{STATIC_ROOT}\\files\\{filename}'
        )
    except ClientError as e:
        raise ClientError(
            f"Failed to download {filename} from S3: {str(e)}") from e


def excel_products(filename, min_col=2, max_col=6, min_row=2):
    """
    Process an Excel file and add products to the database.

    This function checks if the Excel file exists locally, and if not,
    it downloads the file. It then iterates over each sheet in the Excel
    file, matches the sheet names to category names, and adds products
    to the database if they meet certain conditions.

    Args:
        filename (str): The filename of the Excel file.
        min_col (int): The minimum column index to process.
        max_col (int): The maximum column index to process.
        min_row (int): The minimum row index to process.

    Returns:
        str: A message indicating the number of products added to the site.

    """
    if not os.path.exists(f'{STATIC_ROOT}\\files\\{filename}'):
        download_excel(filename)

    wb = openpyxl.load_workbook(f'{STATIC_ROOT}\\files\\{filename}')
    products_cnt = 0

    for page in wb.sheetnames:
        if page not in category_dict:
            continue

        ws = wb[page]
        category = Category.objects.filter(name=category_dict[page])

        if category.exists():
            for row in ws.iter_rows(
                    min_col=min_col, max_col=max_col,
                    min_row=min_row, values_only=True):
                category_id, name, price, count, used_quantity = row

                if (isinstance(id, str) and isinstance(price, (int, float)) and
                        isinstance(count, int) and
                        isinstance(used_quantity, int)):
                    if (category_id and count and len(str(category_id)) == 11
                            and count > 0):
                        if not Product.objects.filter(id=id).exists():
                            Product.objects.create(
                                category=category.first(),
                                id=id,
                                slug=slugify(f'{id}-{name}'),
                                name=name,
                                price=price,
                                count=count,
                                used_quantity=used_quantity
                            )
                            products_cnt += 1

    return f'{products_cnt} products have been added to the site'
